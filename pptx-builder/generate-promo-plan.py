#!/usr/bin/env python3
"""
Ascenda Pipeline — A&P Marketing Budget Tracker Generator

Creates a per-brand A&P Promo Plan spreadsheet matching the master Promo Sheet template.
Pre-populates with category-aware line items and proper Excel formulas.

Usage:
  python3 generate-promo-plan.py --brand "BrandName" --category "Food & Beverage" --output "output.xlsx"
  python3 generate-promo-plan.py --from-json pipeline-data.json
  python3 generate-promo-plan.py --demo
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Category-aware default line items ────────────────────────────────
CATEGORY_PRESETS = {
    "Food & Beverage": {
        "advertising": [
            ("In-Store Sampling Materials", 500),
            ("Influencer / Food Blogger Partnerships", 2000),
            ("Product Photography & Video", 1500),
            ("Press Kits & Mailers", 300),
            ("Sam's Club Co-Op Signage", 0),
        ],
        "promotional": [
            ("Sam's Club Roadshow Promo", 0),
            ("Digital Coupons / Instant Savings", 0),
            ("Google (PMAX, Search, YouTube)", 0),
            ("Meta (Facebook/Instagram)", 0),
            ("Instacart / Retailer Media", 0),
            ("UGC / Affiliate", 1000),
        ],
        "non_working": [
            ("Shopify / E-Commerce Platform", 500),
            ("Email Marketing (Klaviyo/Mailchimp)", 400),
            ("Food Safety Compliance Costs", 200),
            ("Analytics & Reporting Tools", 150),
        ],
    },
    "Wellness/Supplements": {
        "advertising": [
            ("In-Store Demo Materials", 500),
            ("Influencer / Wellness Creator Partnerships", 3000),
            ("Product Photography & Video", 1500),
            ("Press Kits & Mailers", 300),
            ("Sam's Club Co-Op Signage", 0),
        ],
        "promotional": [
            ("Sam's Club Roadshow Promo", 0),
            ("Digital Coupons / Instant Savings", 0),
            ("Google (PMAX, Search, YouTube)", 0),
            ("Meta (Facebook/Instagram)", 0),
            ("Amazon / Retailer Media", 0),
            ("UGC / Affiliate", 1000),
        ],
        "non_working": [
            ("Shopify / E-Commerce Platform", 500),
            ("Email Marketing (Klaviyo/Mailchimp)", 400),
            ("Compliance & Testing Fees", 300),
            ("Analytics & Reporting Tools", 150),
        ],
    },
    "default": {
        "advertising": [
            ("Press Kits & Mailers", 300),
            ("Influencer Partnerships", 3000),
            ("Product Samples / Demos", 0),
            ("Product Photography & Video", 1000),
            ("Sam's Club Co-Op Signage", 0),
        ],
        "promotional": [
            ("UGC / Affiliate", 1000),
            ("Google (PMAX, Search, YouTube)", 0),
            ("Meta (Facebook/Instagram)", 0),
            ("Advertorial (news/feature ads)", 0),
            ("Retailer Media Buys", 0),
            ("Instacart", 0),
        ],
        "non_working": [
            ("Shopify / E-Commerce Platform", 500),
            ("Email Marketing (Klaviyo/Mailchimp)", 1200),
            ("Analytics & Reporting Tools", 0),
        ],
    },
}

# Style constants
NAVY = "0A3161"
BLUE = "0060A9"
GOLD = "D97706"
LIGHT_BLUE = "D6EAF8"
LIGHT_GOLD = "FFF3CD"
LIGHT_GREEN = "D5F5E3"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F3F4"
DARK_GRAY = "2C3E50"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
SECTION_FONT = Font(name="Arial", bold=True, color=NAVY, size=12)
LABEL_FONT = Font(name="Arial", bold=True, color=DARK_GRAY, size=10)
DATA_FONT = Font(name="Arial", color=DARK_GRAY, size=10)
TOTAL_FONT = Font(name="Arial", bold=True, color=NAVY, size=11)
FORMULA_FONT = Font(name="Arial", bold=True, color="000000", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
MONTH_COLS = list(range(5, 17))  # E=5 through P=16


def get_preset(category):
    cat = (category or "").lower()
    if "food" in cat or "beverage" in cat:
        return CATEGORY_PRESETS["Food & Beverage"]
    elif "supplement" in cat or "wellness" in cat:
        return CATEGORY_PRESETS["Wellness/Supplements"]
    return CATEGORY_PRESETS["default"]


def build_promo_plan(brand_name, category, revenue_tier="", retailer="Sam's Club",
                     event_count=0, output_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "A&P Budget Tracker"

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 22
    for col in range(5, 17):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions["Q"].width = 14
    ws.column_dimensions["R"].width = 12

    preset = get_preset(category)
    row = 1

    # ── Title Block ──
    ws.merge_cells("C1:P1")
    ws["C1"] = f"A&P MARKETING BUDGET TRACKER"
    ws["C1"].font = Font(name="Arial", bold=True, color=NAVY, size=16)
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("C2:P2")
    ws["C2"] = f"{brand_name} — {retailer} Roadshow 2026"
    ws["C2"].font = Font(name="Arial", bold=False, color=BLUE, size=12)

    ws.merge_cells("C3:P3")
    ws["C3"] = f"Category: {category}  |  Revenue Tier: {revenue_tier}  |  Prepared by Ascenda Group"
    ws["C3"].font = Font(name="Arial", color="7F8C8D", size=9)

    row = 5

    sections = [
        ("ADVERTISING", "Awareness / Brand Building Efforts", 0.55, preset["advertising"], LIGHT_BLUE),
        ("PROMOTIONAL", "Conversion / Performance Efforts", 0.17, preset["promotional"], LIGHT_GOLD),
        ("NON WORKING", "Subscriptions / Platforms / Consulting Fees", 0.28, preset["non_working"], LIGHT_GREEN),
    ]

    section_total_rows = []
    all_item_rows = []

    for section_name, section_desc, alloc_pct, items, bg_color in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=section_name).font = SECTION_FONT
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=bg_color)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=16)
        ws.cell(row=row, column=5, value=section_desc).font = Font(name="Arial", italic=True, color="7F8C8D", size=9)
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=bg_color)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Column headers
        ws.cell(row=row, column=3, value="Description").font = LABEL_FONT
        ws.cell(row=row, column=4, value="ACCOUNT CODE").font = LABEL_FONT
        for i, month in enumerate(MONTHS):
            cell = ws.cell(row=row, column=MONTH_COLS[i], value=month)
            cell.font = LABEL_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.border = THIN_BORDER
        ws.cell(row=row, column=17, value=f"{alloc_pct:.0%}").font = LABEL_FONT
        ws.cell(row=row, column=17).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=18, value="TOTAL").font = LABEL_FONT
        ws.cell(row=row, column=18).alignment = Alignment(horizontal="center")
        row += 1

        # Data rows (items + empty rows to fill to 8 total)
        item_start_row = row
        total_items = max(len(items), 8)

        for idx in range(total_items):
            desc = items[idx][0] if idx < len(items) else ""
            monthly_val = items[idx][1] if idx < len(items) else 0

            ws.cell(row=row, column=3, value=desc).font = DATA_FONT
            ws.cell(row=row, column=4, value="").font = DATA_FONT  # Account code placeholder

            for col in MONTH_COLS:
                cell = ws.cell(row=row, column=col, value=monthly_val if monthly_val > 0 else 0)
                cell.font = Font(name="Arial", color="0000FF" if monthly_val > 0 else "999999", size=10)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER

            # Row total formula: =SUM(E{row}:P{row})
            total_cell = ws.cell(row=row, column=17)
            total_cell.value = f"=SUM(E{row}:P{row})"
            total_cell.font = FORMULA_FONT
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = Alignment(horizontal="center")

            all_item_rows.append(row)
            row += 1

        item_end_row = row - 1

        # Spacer
        row += 1

        # Section total
        ws.cell(row=row, column=16, value=f"{section_name} TOTAL:").font = TOTAL_FONT
        ws.cell(row=row, column=16).alignment = Alignment(horizontal="right")
        total_cell = ws.cell(row=row, column=17)
        total_cell.value = f"=SUM(Q{item_start_row}:Q{item_end_row})"
        total_cell.font = Font(name="Arial", bold=True, color=NAVY, size=12)
        total_cell.number_format = '$#,##0.00'
        total_cell.alignment = Alignment(horizontal="center")

        section_total_rows.append(row)
        row += 2

    # ── Grand Total Row ──
    ws.cell(row=row, column=4, value="TOTAL").font = Font(name="Arial", bold=True, color=NAVY, size=14)

    # Monthly grand totals
    for col in MONTH_COLS:
        col_letter = get_column_letter(col)
        refs = "+".join([f"{col_letter}{r}" for r in all_item_rows])
        cell = ws.cell(row=row, column=col)
        cell.value = f"=SUM({col_letter}{all_item_rows[0]}:{col_letter}{all_item_rows[-1]})"
        cell.font = TOTAL_FONT
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # Grand total
    grand_total_cell = ws.cell(row=row, column=18)
    section_refs = "+".join([f"Q{r}" for r in section_total_rows])
    grand_total_cell.value = f"={section_refs}"
    grand_total_cell.font = Font(name="Arial", bold=True, color=NAVY, size=14)
    grand_total_cell.number_format = '$#,##0.00'
    grand_total_cell.alignment = Alignment(horizontal="center")

    # ── Summary Section ──
    row += 3
    ws.cell(row=row, column=3, value="BUDGET ALLOCATION SUMMARY").font = Font(name="Arial", bold=True, color=NAVY, size=12)
    row += 1

    summary_items = [
        ("Advertising (Brand Awareness)", f"=Q{section_total_rows[0]}", f"=Q{section_total_rows[0]}/{grand_total_cell.coordinate}"),
        ("Promotional (Performance)", f"=Q{section_total_rows[1]}", f"=Q{section_total_rows[1]}/{grand_total_cell.coordinate}"),
        ("Non-Working (Tools & Platforms)", f"=Q{section_total_rows[2]}", f"=Q{section_total_rows[2]}/{grand_total_cell.coordinate}"),
    ]

    ws.cell(row=row, column=3, value="Category").font = LABEL_FONT
    ws.cell(row=row, column=5, value="Annual Spend").font = LABEL_FONT
    ws.cell(row=row, column=7, value="% of Total").font = LABEL_FONT
    row += 1

    for label, spend_formula, pct_formula in summary_items:
        ws.cell(row=row, column=3, value=label).font = DATA_FONT
        cell_spend = ws.cell(row=row, column=5, value=spend_formula)
        cell_spend.font = FORMULA_FONT
        cell_spend.number_format = '$#,##0'
        cell_pct = ws.cell(row=row, column=7, value=pct_formula)
        cell_pct.font = FORMULA_FONT
        cell_pct.number_format = '0.0%'
        row += 1

    # Total row in summary
    ws.cell(row=row, column=3, value="TOTAL A&P BUDGET").font = TOTAL_FONT
    cell = ws.cell(row=row, column=5, value=f"={grand_total_cell.coordinate}")
    cell.font = Font(name="Arial", bold=True, color=NAVY, size=12)
    cell.number_format = '$#,##0'

    # ── Roadshow Event Overlay Section ──
    row += 3
    ws.cell(row=row, column=3, value="ROADSHOW EVENT COST OVERLAY").font = Font(name="Arial", bold=True, color=NAVY, size=12)
    row += 1
    ws.cell(row=row, column=3, value="(Track per-event costs alongside marketing spend)").font = Font(name="Arial", italic=True, color="7F8C8D", size=9)
    row += 1

    event_headers = ["Event #", "Club/Location", "Date", "Show Length", "Staff Cost",
                     "Product Cost", "Shipping", "Display/Setup", "Total Event Cost", "Revenue", "ROI"]
    for i, header in enumerate(event_headers):
        cell = ws.cell(row=row, column=3 + i, value=header)
        cell.font = LABEL_FONT
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # 10 blank event rows with formulas
    for evt in range(1, 11):
        ws.cell(row=row, column=3, value=evt).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        for col in range(4, 14):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = DATA_FONT

        # Total Event Cost formula (sum of Staff + Product + Shipping + Display)
        ws.cell(row=row, column=11, value=f"=SUM(G{row}:J{row})")
        ws.cell(row=row, column=11).font = FORMULA_FONT
        ws.cell(row=row, column=11).number_format = '$#,##0'

        # ROI formula: (Revenue - Total Cost) / Total Cost
        ws.cell(row=row, column=13, value=f'=IF(K{row}>0,(L{row}-K{row})/K{row},0)')
        ws.cell(row=row, column=13).font = FORMULA_FONT
        ws.cell(row=row, column=13).number_format = '0.0%'

        row += 1

    # Save
    if not output_path:
        safe_name = brand_name.replace(" ", "_").replace("/", "_")
        output_path = f"Ascenda_{safe_name}_A&P_Promo_Plan.xlsx"

    wb.save(output_path)
    print(f"A&P Promo Plan saved: {output_path}")
    print(f"  Brand: {brand_name}")
    print(f"  Category: {category}")
    print(f"  Sections: Advertising ({len(preset['advertising'])} items), "
          f"Promotional ({len(preset['promotional'])} items), "
          f"Non-Working ({len(preset['non_working'])} items)")
    print(f"  Roadshow Event Tracker: 10 event rows with ROI formulas")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate A&P Promo Plan from pipeline data")
    parser.add_argument("--brand", help="Brand name")
    parser.add_argument("--category", help="Product category")
    parser.add_argument("--revenue", help="Revenue tier", default="")
    parser.add_argument("--retailer", help="Target retailer", default="Sam's Club")
    parser.add_argument("--output", help="Output xlsx path")
    parser.add_argument("--from-json", help="Pipeline JSON file")
    parser.add_argument("--demo", action="store_true", help="Generate demo")
    args = parser.parse_args()

    if args.demo:
        build_promo_plan(
            brand_name="GreenLife Supplements",
            category="Wellness/Supplements",
            revenue_tier="$5M - $20M",
            retailer="Sam's Club",
            output_path="Demo_A&P_Promo_Plan.xlsx"
        )
    elif args.from_json:
        with open(args.from_json) as f:
            data = json.load(f)
        build_promo_plan(
            brand_name=data.get("brandName", data.get("company", "Brand")),
            category=data.get("category", ""),
            revenue_tier=data.get("revenue", ""),
            retailer=data.get("retailer", "Sam's Club"),
            output_path=args.output
        )
    elif args.brand:
        build_promo_plan(
            brand_name=args.brand,
            category=args.category or "",
            revenue_tier=args.revenue,
            retailer=args.retailer,
            output_path=args.output
        )
    else:
        parser.print_help()
